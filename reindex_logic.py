"""
================================================================================
  Upgrade Assistant — Index & Data-Stream Reindexing Logic
================================================================================
  UI-agnostic engine for the "♻️ Upgrade Assistant (8.19 - 9.x, Bulk
  Reindexing)" tab in app.py. Wraps the Kibana Upgrade Assistant REST APIs
  (target version 9.x) plus a few read-only Elasticsearch calls used for
  discovery and post-reindex validation.

  Reindex flow performed by the Upgrade Assistant for each index:
      1. Reindex event created                     (UA)
      2. Set index to read-only                    (UA step 20)
      3. Create reindexed-v8-<index> + copy settings (UA step 30)
      4. Reindex documents                         (UA step 40/50)
      5. Alias swap + delete original index        (UA step 60)
      6. Re-point ILM / data-stream aliases        (UA step 70)

  Data streams are upgraded in-place via Elasticsearch's migration reindex
  API (POST /_migration/reindex, a background task): an old write index is
  rolled over first, then each old backing index is reindexed, swapped into
  the stream as `.migrated-…` and the original deleted.

  This module never imports streamlit — the UI lives in app.py.
================================================================================
"""

import io
import re
import threading
import time
from datetime import datetime, timezone

import requests
from requests.utils import quote
import urllib.parse


# An index is considered "ready" for 9.x when it was created at or after 8.0.0.
# version.created is stored as an integer like 7130199 (7.13.1).
ELASTICSEARCH_VERSION_8 = 8000000
TARGET_UPGRADE_VERSION = "9.0.0"

# lastCompletedStep → human-readable label (Kibana Upgrade Assistant)
STEP_CODES = {
    0: "Reindex event created",
    10: "Index group services stopped",
    20: "Set index to read-only",
    30: "Create reindexed index (reindexed-v8-) + copy settings",
    40: "Reindex documents in progress",
    50: "Reindex documents completed",
    60: "Alias swap + delete original index",
    70: "Re-point ILM / data-stream aliases",
}

# reindexOp.status → human-readable label
EVENT_STATUS = {
    0: "In progress",
    1: "Completed",
    2: "Failed",
    3: "Paused",
    4: "Cancelled",
}

# Terminal (non-success) reindexOp statuses
TERMINAL_FAILED = (2, 3, 4)


def make_session(username: str, password: str, verify_ssl: bool = True) -> requests.Session:
    """Basic-auth session shared by every call in this module."""
    session = requests.Session()
    session.auth = (username, password)
    session.verify = verify_ssl
    session.headers.update({"Content-Type": "application/json", "kbn-xsrf": "true"})
    if not verify_ssl:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    return session


def _truncate(text, limit: int = 600) -> str:
    """Shorten a value for the API audit log without hiding the important part."""
    try:
        s = str(text)
    except Exception:
        return "<unprintable>"
    s = " ".join(s.split())
    return s if len(s) <= limit else s[: limit - 3] + "..."


def _redact_url(url: str) -> str:
    """Strip any user:password@ from a URL before logging it."""
    try:
        parts = list(urllib.parse.urlsplit(url))
        parts[1] = parts[1].rsplit("@", 1)[-1] if "@" in parts[1] else parts[1]
        return urllib.parse.urlunsplit(parts)
    except Exception:
        return url


class _LoggingSession:
    """Thin proxy over a requests.Session that records every request/response.

    The worker wraps its real session with this so the audit table can show a
    request + response row for every API call. Responses pass through untouched
    and exceptions still propagate to the caller.
    """

    def __init__(self, inner: requests.Session, record):
        self._inner = inner
        self._record = record

    def _call(self, method: str, url: str, kwargs: dict):
        body = kwargs.get("json")
        if body is None:
            body = kwargs.get("data")
        entry = {
            "ts": now_iso(),
            "method": method,
            "url": _redact_url(url),
            "req_body": _truncate(body, 200),
            "status": None,
            "resp_body": "",
            "ms": None,
        }
        t0 = time.time()
        try:
            resp = self._inner.request(method, url, **kwargs)
        except Exception as exc:
            entry["status"] = "ERR"
            entry["resp_body"] = _truncate(str(exc), 400)
            entry["ms"] = round((time.time() - t0) * 1000)
            self._record(entry)
            raise
        entry["status"] = resp.status_code
        try:
            entry["resp_body"] = _truncate(resp.text, 600)
        except Exception:
            entry["resp_body"] = ""
        entry["ms"] = round((time.time() - t0) * 1000)
        self._record(entry)
        return resp

    def get(self, url, **kwargs):
        return self._call("GET", url, kwargs)

    def post(self, url, **kwargs):
        return self._call("POST", url, kwargs)

    def delete(self, url, **kwargs):
        return self._call("DELETE", url, kwargs)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def needs_reindex(version_created) -> bool:
    return version_created is not None and version_created < ELASTICSEARCH_VERSION_8


def step_code_to_name(step) -> str:
    if step is None:
        return "Preparing reindex"
    return STEP_CODES.get(step, f"Step {step}")


# The 6 user-facing reindex steps (as shown in the Kibana Upgrade Assistant UI).
# Each tuple is (start, done) thresholds on the UA `lastCompletedStep` code plus
# the label template ({index} / {new} are filled in at render time).
REINDEX_STEPS = [
    (20, 30, "Setting {index} index to read-only"),
    (30, 40, "Create {new} index"),
    (40, 50, "Reindex documents"),
    (30, 50, "Copy original index settings from {index} to {new}"),
    (50, 60, "Create {index} alias for {new} index"),
    (60, 60, "Delete original {index} index"),
]


def expected_new_index(index_name: str) -> str:
    """Upgrade Assistant destination naming convention."""
    return f"reindexed-v8-{index_name}"


def build_steps(rec: dict) -> list:
    """
    Resolve the 6-step reindex checklist for a task record.

    `rec["step"]` is the UA lastCompletedStep code; `rec["status"]` is the
    worker status string. Returns [{"label": str, "state": str}] where state is
    one of completed / in_progress / pending / failed. Exactly one step is
    in_progress while the task is running; on failure the first unfinished step
    is marked failed and the rest pending.
    """
    name = rec.get("name") or ""
    new = rec.get("new_index") or expected_new_index(name)
    last_step = _to_int(rec.get("step")) or 0
    status = rec.get("status") or "Queued"
    failed = status in ("Failed", "Cancelled", "Stopped")
    next_up = False
    steps = []
    for start, done, label in REINDEX_STEPS:
        text = label.format(index=name, new=new)
        if status == "Completed" or last_step >= done:
            steps.append({"label": text, "state": "completed"})
        elif failed:
            if not next_up:
                steps.append({"label": text, "state": "failed"})
                next_up = True
            else:
                steps.append({"label": text, "state": "pending"})
        elif last_step >= start and not next_up:
            steps.append({"label": text, "state": "in_progress"})
            next_up = True
        else:
            steps.append({"label": text, "state": "pending"})
    return steps


def _step_milestone_message(index_name: str, new_index, step_code) -> str:
    """Audit-log message for a just-completed UA step (user-facing wording)."""
    new = new_index or expected_new_index(index_name)
    label = None
    for start, done, tpl in REINDEX_STEPS:
        if (step_code or 0) >= done:
            label = tpl
    if label:
        return label.format(index=index_name, new=new)
    return step_code_to_name(step_code)


def _to_int(value):
    if value is None:
        return None
    try:
        if isinstance(value, bool):
            return None
        return int(str(value))
    except (TypeError, ValueError):
        return None


def _dig(key, obj):
    """Recursively find `key` inside nested dicts/lists."""
    if isinstance(obj, dict):
        if key in obj:
            return obj[key]
        for value in obj.values():
            found = _dig(key, value)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = _dig(key, value)
            if found is not None:
                return found
    return None


def _safe_json(resp) -> object:
    """Return the parsed JSON, or None if the body is not valid JSON."""
    try:
        return resp.json()
    except Exception:
        return None


def _version_map(session: requests.Session, es_url: str) -> dict:
    """index name → version.created (int). Empty dict if the call fails.

    Uses the NON-flat settings response with filter_path=**.settings.index.version.created:
    flat_settings + filter_path cannot match flat keys (e.g. `index.version.created`),
    which made earlier versions always return {}. `expand_wildcards=all` ensures hidden
    data-stream backing indices (.ds-*) are included.
    """
    es_url = es_url.rstrip("/")
    vmap = {}
    try:
        resp = session.get(
            f"{es_url}/_all/_settings",
            params={"flat_settings": "false", "expand_wildcards": "all",
                    "filter_path": "**.settings.index.version.created"},
            timeout=60,
        )
        if resp.ok:
            body = _safe_json(resp)
            if isinstance(body, dict):
                for idx, payload in body.items():
                    st = payload.get("settings") or {}
                    created = ((st.get("index") or {}).get("version") or {}).get("created")
                    if created is None:
                        created = st.get("index.version.created")
                    if created is not None:
                        vmap[idx] = _to_int(created)
    except Exception:
        pass
    return vmap


def _parse_deprecation_version(details: str):
    """
    Extract an int version (same encoding as index.version.created) from a
    deprecation detail like 'This index has version: 7.11.1'.

    ES encodes released versions as major*1000000 + minor*10000 + rev*100 + 99
    (e.g. 7.11.1 → 7110199), so the same scheme is used here.
    """
    m = re.search(r"(\d+)\.(\d+)\.(\d+)", details or "")
    if not m:
        return None
    major, minor, revision = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return major * 1000000 + minor * 10000 + revision * 100 + 99


def _int_to_version_str(v) -> str:
    """Decode an ES version.created int back to x.y.z (e.g. 7110199 → 7.11.1)."""
    if v is None:
        return None
    major = v // 1000000
    minor = (v % 1000000) // 10000
    revision = (v % 10000) // 100
    return f"{major}.{minor}.{revision}"


def get_reindex_deprecations(session: requests.Session, es_url: str) -> dict:
    """
    Parse GET /_migration/deprecations for indices that must be reindexed
    before 9.x — the same source the Kibana Upgrade Assistant uses, so the
    result matches the Kibana UI.

    Returns {index_name: {"version": int|None, "version_str": str|None,
                          "reindex_required": bool, "messages": [...]}}
    Only indices flagged reindex-required are returned. Empty dict on any
    failure (callers fall back to version.created detection).
    """
    es_url = es_url.rstrip("/")
    result = {}
    try:
        resp = session.get(f"{es_url}/_migration/deprecations", timeout=60)
        if resp.status_code >= 400:
            return result
        body = _safe_json(resp)
        if not isinstance(body, dict):
            return result
    except Exception:
        return result

    index_settings = body.get("index_settings")
    if not isinstance(index_settings, dict):
        return result

    for idx, entries in index_settings.items():
        if not isinstance(entries, list):
            continue
        version_int = None
        reindex_required = False
        messages = []
        for e in entries:
            if not isinstance(e, dict):
                continue
            msg = e.get("message") or ""
            if msg:
                messages.append(msg)
            meta = e.get("_meta") or {}
            required = bool(meta.get("reindex_required"))
            if not required and e.get("level") == "critical" and "compatibility version" in msg:
                required = True
            if required:
                reindex_required = True
            if version_int is None:
                version_int = _parse_deprecation_version(e.get("details") or "")
        if reindex_required:
            result[idx] = {
                "version": version_int,
                "version_str": _int_to_version_str(version_int),
                "reindex_required": True,
                "messages": messages,
            }
    return result


# ── Discovery (read-only) ────────────────────────────────────────────────────

def get_upgrade_status(session: requests.Session, kibana_url: str,
                       target_version: str = TARGET_UPGRADE_VERSION) -> dict:
    """Kibana Upgrade Assistant readiness for the target upgrade version."""
    kibana_url = kibana_url.rstrip("/")
    resp = session.get(
        f"{kibana_url}/api/upgrade_assistant/status",
        params={"targetVersion": target_version},
        timeout=30,
    )
    if resp.status_code >= 400:
        return {"ready": False, "error": resp.text, "body": None, "counts": {}}
    body = _safe_json(resp)
    if not isinstance(body, dict):
        return {"ready": False, "error": f"Unexpected response from status API: {body!r}",
                "body": body, "counts": {}}
    details = body.get("details")
    counts = {}
    if isinstance(details, dict):
        counts = {k: (len(v) if isinstance(v, list) else v) for k, v in details.items()}
    return {"ready": bool(body.get("readyForUpgrade")), "body": body, "counts": counts}


def discover_indices(session: requests.Session, es_url: str,
                     include_system: bool = False, deprecations: dict = None,
                     stream_map: dict = None, version_map: dict = None) -> list:
    """
    Every index (and data-stream backing index) with its version.created.

    Each record: {name, state, docs_count, store_size, stream, version_created,
                  needs_reindex, is_system, is_closed, can_reindex}

    needs_reindex comes from the /_migration/deprecations reindex-required set
    (authoritative, matches the Kibana UA) OR falls back to
    version.created < 8000000 when deprecations are unavailable.

    `.ds-*` data-stream backing indices are treated as non-system (they are
    user data and the Kibana UA counts them), so they are listed and counted
    even with include_system=False. `stream_map` maps backing index name → the
    data stream that owns it (for the Stream column).
    """
    es_url = es_url.rstrip("/")
    cat = session.get(
        f"{es_url}/_cat/indices",
        params={"format": "json", "expand_wildcards": "open,closed,hidden",
                "h": "index,status,docs.count,store.size,data.stream", "s": "index"},
        timeout=30,
    )
    if cat.status_code >= 400:
        raise RuntimeError(f"GET _cat/indices failed: {cat.text}")
    rows = _safe_json(cat)
    if not isinstance(rows, list):
        rows = []
    vmap = version_map if isinstance(version_map, dict) else _version_map(session, es_url)
    dep = deprecations if isinstance(deprecations, dict) else get_reindex_deprecations(session, es_url)
    stream_map = stream_map or {}

    indices = []
    for row in rows:
        name = row.get("index") or ""
        is_system = name.startswith(".") and not name.startswith(".ds-")
        if is_system and not include_system:
            continue
        state = row.get("status") or "open"
        version_created = vmap.get(name)
        version_int = _to_int(version_created)
        dep_entry = dep.get(name)
        idx_needs = bool(dep_entry) or needs_reindex(version_int)
        disp_version = version_int
        if disp_version is None and dep_entry:
            disp_version = dep_entry.get("version")
        indices.append({
            "name": name,
            "state": state,
            "docs_count": _to_int(row.get("docs.count")),
            "store_size": row.get("store.size", ""),
            "stream": row.get("data.stream") or stream_map.get(name) or None,
            "version_created": disp_version,
            "needs_reindex": idx_needs,
            "is_system": is_system,
            "is_closed": state == "close",
            "can_reindex": bool(idx_needs and state != "close"),
        })

    # Safety net: any index the deprecation API flags but _cat did not return.
    known = {r["name"] for r in indices}
    for name in dep:
        if name in known:
            continue
        is_system = name.startswith(".") and not name.startswith(".ds-")
        if is_system and not include_system:
            continue
        dep_entry = dep[name]
        indices.append({
            "name": name,
            "state": "close",
            "docs_count": None,
            "store_size": "",
            "stream": stream_map.get(name),
            "version_created": dep_entry.get("version"),
            "needs_reindex": True,
            "is_system": is_system,
            "is_closed": True,
            "can_reindex": False,
        })
    return indices


def discover_data_streams(session: requests.Session, es_url: str,
                          deprecations: dict = None,
                          version_map: dict = None) -> dict:
    """
    Data streams → {name, backing_indices, old_backing, write_index,
                    needs_reindex, generation, template, ilm_policy, status}.

    The write index is the last (newest) backing index, per the ES data-stream
    convention. old_backing lists backing indices flagged reindex-required by
    /_migration/deprecations (falling back to version.created < 8.0.0).
    """
    es_url = es_url.rstrip("/")
    resp = session.get(f"{es_url}/_data_stream", timeout=30)
    if resp.status_code >= 400:
        raise RuntimeError(f"GET /_data_stream failed: {resp.text}")
    body = _safe_json(resp)
    if not isinstance(body, dict):
        body = {}
    vmap = version_map if isinstance(version_map, dict) else _version_map(session, es_url)
    dep = deprecations if isinstance(deprecations, dict) else get_reindex_deprecations(session, es_url)

    def _is_old(b):
        if b in dep:
            return True
        ver = _to_int(vmap.get(b))
        return ver is not None and needs_reindex(ver)

    streams = {}
    for ds in body.get("data_streams", []):
        name = ds.get("name")
        backing = [i.get("index_name") for i in ds.get("indices", [])]
        write = backing[-1] if backing else None
        old_backing = [b for b in backing if _is_old(b)]
        streams[name] = {
            "name": name,
            "backing_indices": backing,
            "old_backing": old_backing,
            "write_index": write,
            "needs_reindex": bool(old_backing),
            "generation": ds.get("generation"),
            "template": ds.get("template"),
            "ilm_policy": ds.get("ilm_policy"),
            "status": ds.get("status"),
        }
    return streams


def is_es_endpoint(session: requests.Session, url: str) -> tuple:
    """
    Confirm `url` is an Elasticsearch node, not Kibana or a random proxy.

    Returns (ok: bool, hint: str). Elasticsearch answers `GET /` with JSON that
    includes `cluster_name` and `version.lucene_version`; Kibana's root response
    has neither, so this catches the common "pasted the Kibana URL into the ES
    field" mistake.
    """
    url = url.rstrip("/")
    try:
        resp = session.get(f"{url}/", timeout=20)
        if resp.status_code >= 400:
            return False, (f"HTTP {resp.status_code} on `{url}/` — not an "
                           "Elasticsearch endpoint")
        body = _safe_json(resp)
        if not isinstance(body, dict):
            return False, "Response is not JSON — this does not look like Elasticsearch"
        version = body.get("version") or {}
        if body.get("cluster_name") and version.get("number") and version.get("lucene_version"):
            return True, f"Elasticsearch {version.get('number')}"
        return False, ("This does not look like Elasticsearch (missing cluster_name / "
                       "version.lucene_version). If you entered the Kibana URL, use the "
                       "Elasticsearch endpoint instead (usually port 9200).")
    except Exception as exc:
        return False, str(exc)


def discover_all(session: requests.Session, kibana_url: str, es_url: str,
                 include_system: bool = False,
                 target_version: str = TARGET_UPGRADE_VERSION) -> dict:
    """
    Resilient discovery: run each read independently so one failing endpoint
    (e.g. a `/_data_stream` 404) never blanks the whole scan.

    Returns {"ua_status", "deprecations", "indices", "streams", "warnings"}.
    Whatever succeeded is still usable by the UI, with failures reported as
    human-readable warnings.
    """
    warnings = []
    ua_status = {"ready": False, "error": "Upgrade Assistant status check failed",
                 "body": None, "counts": {}}
    deprecations = {}
    indices = []
    streams = {}

    try:
        ua_status = get_upgrade_status(session, kibana_url, target_version)
        if not ua_status.get("ready") and ua_status.get("error"):
            warnings.append(f"Upgrade Assistant status: {ua_status['error']}")
    except Exception as exc:
        warnings.append(f"Upgrade Assistant status check failed: {exc}")

    try:
        deprecations = get_reindex_deprecations(session, es_url)
    except Exception as exc:
        warnings.append(f"Deprecation scan failed: {exc}")

    try:
        version_map = _version_map(session, es_url)
    except Exception as exc:
        warnings.append(f"Version scan failed: {exc}")
        version_map = {}

    try:
        streams = discover_data_streams(session, es_url, deprecations=deprecations,
                                        version_map=version_map)
    except Exception as exc:
        warnings.append(f"Data-stream discovery failed: {exc}")

    try:
        indices = discover_indices(session, es_url, include_system,
                                   deprecations=deprecations,
                                   stream_map={b: s for s, rec in streams.items()
                                               for b in rec.get("backing_indices", [])},
                                   version_map=version_map)
    except Exception as exc:
        warnings.append(f"Index discovery failed: {exc}")

    return {
        "ua_status": ua_status,
        "deprecations": deprecations,
        "indices": indices,
        "streams": streams,
        "warnings": warnings,
    }


# ── Upgrade Assistant reindex API ────────────────────────────────────────────

def _ua_error_hint(text: str) -> str:
    """Return a short actionable hint for a failed Upgrade Assistant start."""
    low = (text or "").lower()
    if any(k in low for k in ("reindex-required", "not reindex", "is not reindex")):
        return (" The Upgrade Assistant rejected the start — the index may not be "
                "flagged `reindex_required` by its deprecations scan, so the UA "
                "will not reindex it.")
    if "unauthorized" in low or "401" in low or "forbidden" in low or "403" in low:
        return " Check the username/password and the user's API permissions."
    if "not found" in low or "404" in low:
        return (" Check the Kibana URL and that the Upgrade Assistant plugin is "
                "installed/enabled on this Kibana.")
    if "connect" in low or "connection" in low or "timeout" in low or "resolve" in low:
        return " Check the Kibana URL — it could not be reached from this machine."
    return ""


def _unwrap_op(body) -> dict:
    """Kibana's GET reindex-op response nests the op under `reindexOp` while
    the POST response is flat; normalize both to the flat op the callers read."""
    if isinstance(body, dict) and isinstance(body.get("reindexOp"), dict):
        return body["reindexOp"]
    return body if isinstance(body, dict) else {}


def start_reindex(session: requests.Session, kibana_url: str, index_name: str) -> dict:
    kibana_url = kibana_url.rstrip("/")
    resp = session.post(
        f"{kibana_url}/api/upgrade_assistant/reindex/{quote(index_name, safe='')}",
        timeout=30,
    )
    if resp.status_code >= 400:
        return {"_error": resp.text, "status": None}
    return _unwrap_op(resp.json())


def get_reindex_status(session: requests.Session, kibana_url: str, index_name: str) -> dict:
    kibana_url = kibana_url.rstrip("/")
    resp = session.get(
        f"{kibana_url}/api/upgrade_assistant/reindex/{quote(index_name, safe='')}",
        timeout=30,
    )
    if resp.status_code >= 400:
        return {"_error": resp.text, "status": None}
    return _unwrap_op(resp.json())


def get_es_task_status(session: requests.Session, es_url: str, task_id: str):
    """
    Live doc-level status of the underlying ES reindex task via /_tasks.

    `task_id` is the `reindexTaskId` from the Upgrade Assistant reindex op
    (format "nodeId:taskId"). Returns None on 404 / errors so callers never
    fail a poll because the task already finished or the node is unreachable.
    """
    es_url = es_url.rstrip("/")
    try:
        resp = session.get(
            f"{es_url}/_tasks/{quote(task_id, safe=':')}",
            timeout=15,
        )
        if resp.status_code == 404:
            return None
        if resp.status_code >= 400:
            return None
        body = resp.json() or {}
        task = body.get("task") or {}
        st = task.get("status") or {}
        total = _to_int(st.get("total"))
        created = _to_int(st.get("created"))
        updated = _to_int(st.get("updated"))
        deleted = _to_int(st.get("deleted"))
        percent = None
        if total:
            percent = round(((created or 0) + (updated or 0)) / total * 100, 1)
        nanos = _to_int(task.get("running_time_in_nanos"))
        return {
            "completed": bool(body.get("completed")),
            "docs_total": total,
            "docs_created": created,
            "docs_updated": updated,
            "docs_deleted": deleted,
            "percent": percent,
            "running_ms": (nanos / 1_000_000) if nanos else None,
        }
    except Exception:
        return None


def cancel_reindex(session: requests.Session, kibana_url: str, index_name: str) -> dict:
    kibana_url = kibana_url.rstrip("/")
    resp = session.delete(
        f"{kibana_url}/api/upgrade_assistant/reindex/{quote(index_name, safe='')}",
        timeout=30,
    )
    if resp.status_code >= 400:
        return {"_error": resp.text}
    return resp.json() if resp.text else {}


# ── ES migration reindex API (data streams) ──────────────────────────────────
# Data-stream backing indices cannot carry aliases, so the UA alias-based flow
# rejects them. ES instead upgrades a whole stream in the background with the
# migration reindex API; status is polled per data stream name (the task id is
# internal, so there is no task-id round-trip).

def migrate_reindex_stream(session: requests.Session, es_url: str, stream_name: str) -> dict:
    """
    Start ES's migration reindex on a data stream (mode=upgrade, in-place).

    POST /_migration/reindex  {"source": {"index": stream}, "mode": "upgrade"}
    Returns the acknowledged response or {"_error": ...}. Runs as a background
    task tracked per data stream name.
    """
    es_url = es_url.rstrip("/")
    try:
        resp = session.post(
            f"{es_url}/_migration/reindex",
            json={"source": {"index": stream_name}, "mode": "upgrade"},
            timeout=30,
        )
    except Exception as exc:
        return {"_error": str(exc)}
    if resp.status_code >= 400:
        return {"_error": resp.text}
    body = _safe_json(resp)
    return body if isinstance(body, dict) else {}


def get_migrate_reindex_status(session: requests.Session, es_url: str, stream_name: str) -> dict:
    """
    Poll a data-stream migration reindex task.

    Returns the status JSON (complete, successes, pending, in_progress, errors),
    {"_not_found": True} on 404 (no task / status expired), or {"_error": ...}.
    """
    es_url = es_url.rstrip("/")
    try:
        resp = session.get(
            f"{es_url}/_migration/reindex/{quote(stream_name, safe='')}/_status",
            timeout=30,
        )
    except Exception as exc:
        return {"_error": str(exc)}
    if resp.status_code == 404:
        return {"_not_found": True}
    if resp.status_code >= 400:
        return {"_error": resp.text}
    body = _safe_json(resp)
    return body if isinstance(body, dict) else {}


def cancel_migrate_reindex(session: requests.Session, es_url: str, stream_name: str) -> dict:
    """POST /_migration/reindex/{stream}/_cancel — stop a running migration task."""
    es_url = es_url.rstrip("/")
    try:
        resp = session.post(
            f"{es_url}/_migration/reindex/{quote(stream_name, safe='')}/_cancel",
            timeout=30,
        )
    except Exception as exc:
        return {"_error": str(exc)}
    if resp.status_code >= 400:
        return {"_error": resp.text}
    return _safe_json(resp) or {}


def get_data_stream_indices(session: requests.Session, es_url: str, stream_name: str) -> list:
    """Current backing index names of a data stream (empty list on failure)."""
    es_url = es_url.rstrip("/")
    try:
        resp = session.get(
            f"{es_url}/_data_stream/{quote(stream_name, safe='')}",
            timeout=30,
        )
    except Exception:
        return []
    if resp.status_code >= 400:
        return []
    body = _safe_json(resp)
    if not isinstance(body, dict):
        return []
    for ds in body.get("data_streams", []):
        if ds.get("name") == stream_name:
            return [i.get("index_name") for i in ds.get("indices", [])]
    return []


def _migration_new_name(old_index: str) -> str:
    """Migration swaps a backing index in under `.migrated-<original-without-dot>`
    (docs: `.ds-foo-…-000001` → `.migrated-ds-foo-…-000001`)."""
    return f".migrated-{old_index.lstrip('.')}"


def _map_migrated_backing(old_backing: list, current_backing: list) -> dict:
    """Map old backing names → their replacement after a migration, matched on
    the trailing generation number (preserved across the swap + rollover)."""
    mapping = {}
    for old in old_backing:
        gen = old.rsplit("-", 1)[-1]
        matches = [c for c in current_backing
                   if c.rsplit("-", 1)[-1] == gen and c != old]
        mapping[old] = matches[0] if matches else None
    return mapping


def rollover_stream(session: requests.Session, es_url: str, stream_name: str) -> dict:
    """Roll over a data stream's write index (creates a new, current-version index)."""
    es_url = es_url.rstrip("/")
    resp = session.post(
        f"{es_url}/{quote(stream_name, safe='')}/_rollover",
        timeout=60,
    )
    if resp.status_code >= 400:
        return {"_error": resp.text}
    return resp.json() or {}


def validate_index(session: requests.Session, es_url: str, index_name: str) -> dict:
    """
    Post-reindex validation. After a successful UA reindex the original name is
    an alias pointing at the reindexed-v8-… index, so version.created read under
    the original name reflects the new index.
    """
    es_url = es_url.rstrip("/")
    q = quote(index_name, safe="")
    version_created = None
    try:
        vs = session.get(
            f"{es_url}/{q}/_settings",
            params={"flat_settings": "true"},
            timeout=30,
        )
        if vs.ok:
            version_created = _dig("index.version.created", vs.json())
    except Exception:
        pass
    docs_count = None
    try:
        cnt = session.get(f"{es_url}/{q}/_count", timeout=30)
        if cnt.ok:
            docs_count = cnt.json().get("count")
    except Exception:
        pass
    ver_int = _to_int(version_created)
    ok = ver_int is not None and ver_int >= ELASTICSEARCH_VERSION_8
    return {"ok": ok, "index_name": index_name, "version_created": ver_int,
            "docs_count": docs_count}


# ── Data-stream read-only (9.x upgrade prep) ─────────────────────────────────

# Platform-internal data-stream prefixes. Excluded from the "Data Streams
# Read-Only" candidate list unless include_system is True. Only genuinely
# Fleet-managed streams are hidden by default — APM/Synthetics/metrics streams
# carry real application data that the operator legitimately needs to freeze,
# so they are kept visible (they are reported as "system" only via the stats'
# data-stream `system` flag when available).
SYSTEM_STREAMS_PREFIXES = (
    "fleet-", ".fleet-",
)

# Operations sent to the Upgrade Assistant `update_index` endpoint when marking
# a data stream read-only. blockWrite sets index.blocks.write=true; unfreeze is
# needed for legacy (7.x) frozen backing indices, which must be thawed before 9.x.
READ_ONLY_OPERATIONS = ("blockWrite", "unfreeze")

# Batch size for the comma-separated index names in one update_index / _settings call.
_UPDATE_INDEX_BATCH = 50


def _ts_zero(ts) -> bool:
    """True when a data-stream `maximum_timestamp` means "no data written yet".

    `_data_stream/*/_stats` reports `maximum_timestamp` as the number 0 (or the
    string "0") for a stream that has never received documents.
    """
    if isinstance(ts, bool):
        return False
    if isinstance(ts, (int, float)):
        return ts == 0
    if isinstance(ts, str):
        return ts.strip() in ("", "0")
    return False


def get_data_stream_stats(session: requests.Session, es_url: str) -> dict:
    """
    GET _data_stream/*/_stats?human=true → per data stream: store size,
    maximum_timestamp (last updated) and backing-index count.

    `maximum_timestamp` reflects the last time data was written to the stream —
    used to decide which streams are safe to freeze (no longer being updated).
    A value of `0` (or `"0"`) means the stream has never received documents —
    nothing to freeze. NOTE: ES 8.x keys each entry by `data_stream`, ES 7.x by
    `name`; both are read so the same code works across versions.
    """
    es_url = es_url.rstrip("/")
    resp = session.get(
        f"{es_url}/_data_stream/*/_stats",
        params={"human": "true"},
        timeout=30,
    )
    if resp.status_code >= 400:
        raise RuntimeError(f"GET _data_stream/_stats failed: {resp.text}")
    body = _safe_json(resp)
    if not isinstance(body, dict):
        return {}
    stats = {}
    for ds in body.get("data_streams", []):
        name = ds.get("name") or ds.get("data_stream")
        if not name:
            continue
        backing = [i.get("index_name") for i in ds.get("indices", []) if i.get("index_name")]
        stats[name] = {
            "name": name,
            "backing_indices": backing,
            "write_index": backing[-1] if backing else None,
            "generation": ds.get("generation"),
            "store_size": ds.get("store_size"),
            "store_size_bytes": ds.get("store_size_bytes"),
            "maximum_timestamp": ds.get("maximum_timestamp"),
            "docs_count": ds.get("docs_count"),
            "backing_count": _to_int(ds.get("backing_indices")) or (len(backing) if backing else 0),
        }
    return stats


def discover_stream_readonly(session: requests.Session, es_url: str,
                             include_system: bool = False,
                             deprecations: dict = None,
                             version_map: dict = None) -> dict:
    """
    Data streams with the info the read-only sub-tab needs, in one pass:

    - `discover_data_streams()` → backing / old_backing / needs_reindex
    - `get_data_stream_stats()`  → maximum_timestamp / docs / store

    Never raises: each source is guarded and its failure is reported as a
    warning instead (same resilience philosophy as `discover_all`).

    Returns {"records": [...], "excluded_system": [...], "warnings": [...]}.
    `records` is the per-stream inventory (backing-index count = "Backing"
    column, `maximum_timestamp` = "Last updated"). Streams matching
    SYSTEM_STREAMS_PREFIXES (Fleet-internal) are dropped unless include_system
    is True, and listed under `excluded_system` so the UI can tell the operator
    why they are missing.
    """
    warnings = []
    try:
        base = discover_data_streams(session, es_url,
                                     deprecations=deprecations, version_map=version_map)
    except Exception as exc:
        warnings.append(f"Data-stream discovery failed: {exc}")
        base = {}
    try:
        stats = get_data_stream_stats(session, es_url)
    except Exception as exc:
        warnings.append(f"Data-stream stats (`_data_stream/*/_stats`) failed: {exc}")
        stats = {}

    records = []
    excluded = []
    for name, rec in base.items():
        if not include_system and name.startswith(SYSTEM_STREAMS_PREFIXES):
            excluded.append(name)
            continue
        s = stats.get(name, {})
        records.append({
            "name": name,
            "backing_indices": rec.get("backing_indices") or s.get("backing_indices") or [],
            "write_index": rec.get("write_index") or s.get("write_index"),
            "old_backing": rec.get("old_backing") or [],
            "needs_reindex": bool(rec.get("needs_reindex")),
            "maximum_timestamp": s.get("maximum_timestamp"),
            "docs_count": s.get("docs_count"),
            "store_size": s.get("store_size"),
            "backing_count": len(rec.get("backing_indices") or []) or s.get("backing_count") or 0,
            "is_empty": bool(s) and (
                s.get("docs_count") == 0 or _ts_zero(s.get("maximum_timestamp"))
            ),
        })
    return {"records": records, "excluded_system": excluded, "warnings": warnings}


def _extract_http_error(resp, body=None) -> str:
    """Human-readable error from an HTTP response, preferring the `message`
    field (Kibana/ES wrap the real reason there) over the generic `error`."""
    if resp is None:
        return "request failed (no response)"
    if isinstance(body, dict):
        msg = body.get("message") or ""
        err = body.get("error") or ""
        if isinstance(err, dict):
            err_type = err.get("type") or ""
            reason = err.get("reason") or ""
            if reason:
                return f"{err_type}: {reason}" if err_type else str(reason)
            err = err_type or err
        if msg and err:
            return f"{err}: {msg}" if err not in str(msg) else str(msg)
        if msg:
            return str(msg)
        if err:
            return str(err)
    text = resp.text or ""
    if text:
        return _truncate(text, 300)
    return f"HTTP {resp.status_code}"


def _index_readonly_settings(session: requests.Session, es_url: str,
                             names: list) -> dict:
    """
    Current write-block + frozen state per index, read once for a batch:

        GET {es}/{csv}/_settings?filter_path=*.settings.index.blocks.write,*.settings.index.frozen

    Returns {index_name: {"blocks_write": bool, "frozen": bool}}. Indices that
    could not be read default to False (so they are still attempted).
    """
    es_url = es_url.rstrip("/")
    out = {}
    for i in range(0, len(names), _UPDATE_INDEX_BATCH):
        batch = names[i:i + _UPDATE_INDEX_BATCH]
        path_csv = ",".join(quote(n, safe="") for n in batch)
        try:
            resp = session.get(
                f"{es_url}/{path_csv}/_settings",
                params={"filter_path": "*.settings.index.blocks.write,*.settings.index.frozen"},
                timeout=30,
            )
        except Exception:
            for n in batch:
                out[n] = {"blocks_write": False, "frozen": False}
            continue
        body = _safe_json(resp)
        if not isinstance(body, dict):
            for n in batch:
                out[n] = {"blocks_write": False, "frozen": False}
            continue
        for n in batch:
            entry = body.get(n, {})
            st = (entry.get("settings") or {}).get("index") or {}
            out[n] = {
                "blocks_write": _dig("write", st) in (True, "true"),
                "frozen": _dig("frozen", st) in (True, "true"),
            }
    return out


def _ua_update_index(session: requests.Session, kibana_url: str,
                     names: list, ops: tuple) -> list:
    """POST the Upgrade Assistant update_index endpoint for a CSV of names."""
    kibana_url = kibana_url.rstrip("/")
    results = []
    for i in range(0, len(names), _UPDATE_INDEX_BATCH):
        batch = names[i:i + _UPDATE_INDEX_BATCH]
        path_csv = ",".join(quote(n, safe="") for n in batch)
        try:
            resp = session.post(
                f"{kibana_url}/api/upgrade_assistant/update_index/{path_csv}",
                json={"operations": list(ops)},
                timeout=60,
            )
        except Exception as exc:
            resp = None
        body = _safe_json(resp) if resp is not None else None
        ok = resp is not None and resp.status_code < 400
        status = resp.status_code if resp is not None else None
        for n in batch:
            results.append({
                "index": n,
                "http_status": status,
                "ok": ok,
                "body": body if ok else None,
                "error": None if ok else _extract_http_error(resp, body),
                "operations": list(ops),
            })
    return results


def _es_set_read_only(session: requests.Session, es_url: str,
                      index: str, ops: tuple) -> dict:
    """
    Direct Elasticsearch fallback: unfreeze via the `_unfreeze` API when needed
    (7.x legacy frozen indices must be thawed before 9.x), then set
    `index.blocks.write=true` via update settings.
    """
    es_url = es_url.rstrip("/")
    q = quote(index, safe="")
    if "unfreeze" in ops:
        try:
            resp = session.post(f"{es_url}/{q}/_unfreeze", timeout=60)
            body = _safe_json(resp)
            if resp.status_code >= 400:
                return {"index": index, "http_status": resp.status_code, "ok": False,
                        "body": None, "error": _extract_http_error(resp, body),
                        "operations": list(ops)}
        except Exception as exc:
            return {"index": index, "http_status": None, "ok": False, "body": None,
                    "error": str(exc), "operations": list(ops)}
    try:
        resp = session.put(
            f"{es_url}/{q}/_settings",
            json={"index.blocks.write": True},
            timeout=60,
        )
        body = _safe_json(resp)
        ok = resp.status_code < 400
        return {"index": index, "http_status": resp.status_code, "ok": ok,
                "body": body if ok else None,
                "error": None if ok else _extract_http_error(resp, body),
                "operations": list(ops)}
    except Exception as exc:
        return {"index": index, "http_status": None, "ok": False, "body": None,
                "error": str(exc), "operations": list(ops)}


def mark_indices_read_only(session: requests.Session, kibana_url: str, es_url: str,
                           index_names: list, operations: tuple = None,
                           use_ua: bool = True) -> list:
    """
    Make indices read-only. The Upgrade Assistant refuses to block a data
    stream's **write index** (it returns 400 "cannot add a block to the following
    data stream write indices"), so data streams go straight to Elasticsearch
    (`use_ua=False`); regular indices prefer the UA and fall back to ES if the
    UA rejects the request (e.g. older 7.x UAs return 400 for `unfreeze`):

        POST {kibana}/api/upgrade_assistant/update_index/{csv}  {"operations": [...]}
        PUT  {es}/{index}/_settings                             {"index.blocks.write": true}

    Unless `operations` is given, the operations are derived per index from its
    current settings so `unfreeze` is only sent when the index is actually
    frozen (7.x legacy frozen indices must be thawed before 9.x — on an 8.x
    cluster nothing is frozen, so only `blockWrite` is ever sent). Returns one
    result dict per index:

        {"index", "http_status", "ok", "body", "error", "operations", "method"}
    where `method` is "upgrade_assistant" | "es_settings" | "already".
    """
    operations = tuple(operations) if operations else None
    names = [n for n in index_names if n]
    current = _index_readonly_settings(session, es_url, names)
    plan = {}
    for n in names:
        if operations is not None:
            plan[n] = operations
            continue
        st = current.get(n) or {}
        ops = []
        if not st.get("blocks_write"):
            ops.append("blockWrite")
        if st.get("frozen"):
            ops.append("unfreeze")
        plan[n] = tuple(ops)

    groups = {}
    for n in names:
        groups.setdefault(plan[n], []).append(n)

    results = []
    for ops, grp in groups.items():
        if not ops:
            for n in grp:
                results.append({
                    "index": n, "http_status": 200, "ok": True, "body": None,
                    "error": None, "operations": [], "method": "already",
                })
            continue
        if not use_ua:
            for n in grp:
                es_r = _es_set_read_only(session, es_url, n, ops)
                es_r["method"] = "es_settings" if es_r["ok"] else "failed"
                results.append(es_r)
            continue
        ua_results = _ua_update_index(session, kibana_url, grp, ops)
        for r in ua_results:
            if r["ok"]:
                r["method"] = "upgrade_assistant"
                results.append(r)
            else:
                es_r = _es_set_read_only(session, es_url, r["index"], ops)
                es_r["method"] = "es_settings" if es_r["ok"] else "failed"
                if es_r["ok"]:
                    es_r["note"] = f"Upgrade Assistant rejected the request: {r['error']}"
                results.append(es_r)
    return results


def check_indices_blocks_write(session: requests.Session, es_url: str,
                               index_names: list) -> dict:
    """
    Verify read-only status: GET {es}/{csv}/_settings?filter_path=*.settings.index.blocks.write

    Returns {index_name: bool} — True only when `index.blocks.write` is "true".
    Indices missing from the filter_path response are reported False.
    """
    es_url = es_url.rstrip("/")
    out = {}
    names = [n for n in index_names if n]
    for i in range(0, len(names), _UPDATE_INDEX_BATCH):
        batch = names[i:i + _UPDATE_INDEX_BATCH]
        path_csv = ",".join(quote(n, safe="") for n in batch)
        try:
            resp = session.get(
                f"{es_url}/{path_csv}/_settings",
                params={"filter_path": "*.settings.index.blocks.write"},
                timeout=30,
            )
        except Exception:
            for n in batch:
                out[n] = False
            continue
        body = _safe_json(resp)
        if not isinstance(body, dict):
            for n in batch:
                out[n] = False
            continue
        for n in batch:
            entry = body.get(n, {})
            out[n] = _dig("write", entry) is True or _dig("write", entry) == "true"
    return out


def apply_streams_read_only(session: requests.Session, kibana_url: str, es_url: str,
                            stream_records: list, use_ua: bool = False) -> list:
    """
    Mark every backing index of the given data streams read-only, then verify
    `index.blocks.write` on each. Data streams bypass the Upgrade Assistant
    (`use_ua=False` by default) because the UA refuses to block a data stream's
    write index; the block is applied directly via Elasticsearch settings.
    Returns the input records enriched with:

        read_only       → bool (all backing indices verified read-only)
        applied         → bool (update_index call attempted)
        http_statuses   → [int]
        error           → stream-level error text (from any failing batch)
        index_results   → [{index, is_write, http_status, ok, blocks_write}]
    """
    es_url = es_url.rstrip("/")
    enriched = []
    for rec in stream_records:
        out = dict(rec)
        backing = list(rec.get("backing_indices") or [])
        write_index = rec.get("write_index")
        if not backing:
            out.update({"read_only": False, "applied": False, "http_statuses": [],
                        "error": "No backing indices.", "index_results": []})
            enriched.append(out)
            continue
        marks = mark_indices_read_only(session, kibana_url, es_url, backing,
                                       use_ua=use_ua)
        checks = check_indices_blocks_write(session, es_url, backing)
        statuses = [m["http_status"] for m in marks]
        errors = [m["error"] for m in marks if m.get("error")]
        index_results = []
        for m in marks:
            index_results.append({
                "index": m["index"],
                "is_write": m["index"] == write_index,
                "http_status": m["http_status"],
                "ok": m["ok"],
                "blocks_write": checks.get(m["index"], False),
                "method": m.get("method") or "",
                "operations": m.get("operations") or [],
                "note": m.get("note") or "",
            })
        out.update({
            "read_only": bool(index_results) and all(r["blocks_write"] for r in index_results),
            "applied": True,
            "http_statuses": statuses,
            "error": errors[0] if errors else None,
            "index_results": index_results,
        })
        enriched.append(out)
    return enriched


def build_readonly_report_excel(results: list) -> bytes:
    """Two-sheet workbook: Data Streams summary + Backing Indices detail."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws = wb.active
    ws.title = "Data Streams"
    ws.append(["Data Stream", "Backing Indices", "Write Index", "Last Updated",
               "Docs", "Store", "Old Backing", "Needs Reindex", "Read-Only Applied",
               "Fully Verified Read-Only", "Error"])
    for r in results:
        ws.append([
            r.get("name"), r.get("backing_count"), r.get("write_index") or "",
            r.get("maximum_timestamp") or "", r.get("docs_count") or "",
            r.get("store_size") or "", len(r.get("old_backing") or []),
            "✅" if r.get("needs_reindex") else "—",
            "✅" if r.get("applied") else "—",
            "✅" if r.get("read_only") else "❌",
            (r.get("error") or "")[:200],
        ])

    ws2 = wb.create_sheet("Backing Indices")
    ws2.append(["Data Stream", "Backing Index", "Is Write Index", "HTTP Status",
                "Operation OK", "Blocks.Write Verified"])
    for r in results:
        for ir in r.get("index_results", []):
            ws2.append([
                r.get("name"), ir.get("index"),
                "✅" if ir.get("is_write") else "",
                ir.get("http_status"),
                "✅" if ir.get("ok") else "❌",
                "✅" if ir.get("blocks_write") else "❌",
            ])

    for sheet, widths in (
        (ws, [42, 14, 30, 24, 12, 14, 12, 12, 14, 18, 60]),
        (ws2, [42, 48, 12, 12, 12, 20]),
    ):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Background worker ────────────────────────────────────────────────────────

class ReindexWorker:
    """
    Processes a queue of indices / data streams one at a time on a background
    thread. The UI reads `.state` on every rerun; controls (halt / resume /
    cancel / hard stop) are safe to call from the main (Streamlit) thread.

    Halt semantics: the worker finishes the in-flight reindex event, then stops
    picking up new tasks so the user can validate in the middle of the upgrade.
    """

    def __init__(self, kibana_url: str, es_url: str, username: str, password: str,
                 verify_ssl: bool = True, poll_interval: float = 2.0,
                 max_events: int = 2000):
        self.connection_key = f"{kibana_url.strip().rstrip('/')}|{es_url.strip().rstrip('/')}"
        self.kibana_url = kibana_url.rstrip("/")
        self.es_url = es_url.rstrip("/")
        self.username = username
        self.password = password
        self.verify_ssl = verify_ssl
        self.poll_interval = poll_interval
        self.max_events = max_events
        self.pause_after_each = False

        self._lock = threading.Lock()
        self._thread = None
        self._session = None
        self._halt = False
        self._stop = False
        self._running = False
        self._queue = []
        self._tasks = {}
        self._streams = {}
        self._events = []
        self._api_log = []
        self._active = None

    # ── controls (call from the UI thread) ──────────────────────────────────
    def start(self):
        with self._lock:
            if self._thread and self._thread.is_alive():
                return
            self._thread = threading.Thread(target=self._run, daemon=True)
            self._running = True
            self._halt = False
            self._thread.start()

    def halt(self):
        with self._lock:
            self._halt = True

    def resume(self):
        with self._lock:
            self._halt = False

    def cancel_current(self):
        with self._lock:
            name = self._active
        if name:
            self._record_event(name, "cancel-requested", f"Cancel requested for {name}")
            try:
                if name in self._streams:
                    cancel_migrate_reindex(self._session, self.es_url, name)
                else:
                    cancel_reindex(self._session, self.kibana_url, name)
            except Exception as exc:
                self._record_event(name, "cancel-failed", str(exc))

    def hard_stop(self):
        with self._lock:
            self._stop = True
            self._halt = True
        self.cancel_current()
        self._record_event(None, "hard-stop", "Hard stop requested; no new tasks will start.")

    # ── queueing ─────────────────────────────────────────────────────────────
    def enqueue(self, name: str, kind: str = "index", front: bool = False,
                meta: dict = None) -> bool:
        with self._lock:
            existing = self._tasks.get(name)
            if existing and existing["status"] in ("Queued", "Running", "Rolling Over", "In progress"):
                return False
            rec = self._new_record(name, kind, None)
            self._apply_meta(rec, meta)
            self._tasks[name] = rec
            if front:
                self._queue.insert(0, {"kind": kind, "name": name})
            else:
                self._queue.append({"kind": kind, "name": name})
            return True

    @staticmethod
    def _apply_meta(rec: dict, meta: dict):
        """Stamp initial discovery info (version, docs, alias) onto a task record."""
        if not meta:
            return
        if meta.get("version_created") is not None:
            rec["version_created"] = meta.get("version_created")
        if meta.get("docs_count") is not None:
            rec["docs"] = str(meta.get("docs_count"))
        if meta.get("alias"):
            rec["alias"] = meta.get("alias")

    def enqueue_stream(self, stream_name: str, record: dict) -> bool:
        with self._lock:
            existing = self._tasks.get(stream_name)
            if existing and existing["status"] in ("Queued", "Running", "Rolling Over", "In progress"):
                return False
            self._streams[stream_name] = record
            rec = self._new_record(stream_name, "stream", None)
            rec["write_index"] = record.get("write_index")
            rec["old_backing_count"] = len(record.get("old_backing", []))
            rec["backing_count"] = len(record.get("backing_indices", []))
            self._apply_meta(rec, record)
            for b in record.get("backing_indices", []):
                self._tasks.setdefault(b, self._new_record(b, "index", stream_name))
            self._tasks[stream_name] = rec
            self._queue.append({"kind": "stream", "name": stream_name})
            return True

    def reset_queue(self, items: list, stream_records: dict = None,
                    index_meta: dict = None) -> int:
        """
        Replace the queue with `items` (list of (kind, name)).

        Queued-only task records are dropped (history of in-flight / completed /
        failed tasks is preserved for the tracker), then the new items are
        enqueued with dedupe against anything still tracked. Stream items reuse
        the stored stream record (or the caller-supplied `stream_records`).
        `index_meta` maps index name → discovery record for initial info.
        """
        with self._lock:
            for n, r in list(self._tasks.items()):
                if r["status"] == "Queued":
                    del self._tasks[n]
            self._queue = []
        added = 0
        for kind, name in items:
            if kind == "stream":
                rec = (stream_records or {}).get(name) or self._streams.get(name)
                if rec is None:
                    continue
                if self.enqueue_stream(name, rec):
                    added += 1
            else:
                meta = (index_meta or {}).get(name)
                if self.enqueue(name, kind="index", meta=meta):
                    added += 1
        return added

    def _new_record(self, name, kind, stream):
        return {
            "name": name, "kind": kind, "stream": stream,
            "status": "Queued", "step": None, "step_name": None,
            "progress": None, "new_index": None,
            "task_id": None, "task_details": None, "docs": None,
            "alias": name, "version_created": None,
            "started_at": None, "completed_at": None, "error": None, "event": None,
        }

    # ── state snapshot for the UI ───────────────────────────────────────────
    @property
    def state(self) -> dict:
        with self._lock:
            tasks = {n: dict(r) for n, r in self._tasks.items()}
            events = [dict(e) for e in self._events[-500:]]
            api_log = [dict(e) for e in self._api_log[-500:]]
            queue = [dict(t) for t in self._queue]
            running = self._running
            halted = self._halt
            stopped = self._stop
            pause_each = self.pause_after_each
            active = self._active
        statuses = [r.get("status") for r in tasks.values()]
        summary = {
            "queued": statuses.count("Queued"),
            "running": statuses.count("Running") + statuses.count("Rolling Over"),
            "completed": statuses.count("Completed"),
            "failed": statuses.count("Failed"),
            "cancelled": statuses.count("Cancelled") + statuses.count("Stopped"),
        }
        return {
            "running": running,
            "halted": halted,
            "stopped": stopped,
            "pause_after_each": pause_each,
            "current": active,
            "queue": queue,
            "tasks": tasks,
            "events": events,
            "api_log": api_log,
            "summary": summary,
        }

    @property
    def is_alive(self) -> bool:
        t = self._thread
        return t is not None and t.is_alive()

    # ── worker thread ───────────────────────────────────────────────────────
    def _run(self):
        try:
            base = make_session(self.username, self.password, self.verify_ssl)
            self._session = _LoggingSession(base, self._record_api_call)
        except Exception as exc:
            with self._lock:
                self._running = False
            self._record_event(None, "connect-failed", str(exc))
            return
        self._record_event(None, "worker-started", f"Connected: {self.connection_key}")

        while True:
            try:
                with self._lock:
                    if self._stop:
                        self._running = False
                        return
                    if self._queue:
                        task = self._queue.pop(0)
                    else:
                        task = None
                if task is None:
                    time.sleep(self.poll_interval)
                    continue

                kind, name = task.get("kind"), task["name"]
                try:
                    if kind == "stream":
                        self._process_stream(name)
                    else:
                        self._process_index(name)
                except Exception as exc:
                    rec = self._tasks.get(name)
                    if rec:
                        rec["status"] = "Failed"
                        rec["error"] = str(exc)
                        rec["completed_at"] = now_iso()
                    self._record_event(name, "worker-error", str(exc))
                    with self._lock:
                        self._halt = True

                # Wait while halted (manual pause, pause-after-each, or flag-and-stop).
                while True:
                    with self._lock:
                        should_wait = self._halt and not self._stop
                    if not should_wait:
                        break
                    time.sleep(self.poll_interval)
            except Exception as exc:
                with self._lock:
                    self._running = False
                self._record_event(None, "worker-crash", str(exc))
                return

    def _process_index(self, name):
        rec = self._tasks.get(name)
        if rec is None:
            return
        self._active = name
        try:
            ok = self._reindex_one(name)
            if not ok:
                with self._lock:
                    self._halt = True  # flag-and-stop on failure
            elif self.pause_after_each:
                with self._lock:
                    self._halt = True
        finally:
            self._active = None

    def _reindex_one(self, name) -> bool:
        """Full lifecycle for one index; returns True only on completed."""
        rec = self._tasks[name]
        rec["status"] = "Running"
        rec["started_at"] = now_iso()

        try:
            op = start_reindex(self._session, self.kibana_url, name)
        except Exception as exc:
            msg = f"{exc}{_ua_error_hint(str(exc))}"
            rec["status"] = "Failed"
            rec["error"] = msg
            rec["completed_at"] = now_iso()
            self._record_event(name, "start-failed", msg)
            return False

        if op.get("_error"):
            msg = f"{op['_error']}{_ua_error_hint(op['_error'])}"
            rec["status"] = "Failed"
            rec["error"] = msg
            rec["completed_at"] = now_iso()
            self._record_event(name, "start-failed", msg)
            return False

        rec["event"] = op
        if op.get("reindexTaskId"):
            rec["task_id"] = op.get("reindexTaskId")
        if op.get("status") == 1:
            rec["status"] = "Completed"
            rec["step_name"] = "Completed"
            rec["progress"] = 100.0
            rec["new_index"] = op.get("newIndexName")
            rec["completed_at"] = now_iso()
            self._record_event(name, "completed", f"Already reindexed ({rec['new_index']})", op)
            return True
        if op.get("status") in TERMINAL_FAILED:
            msg = op.get("errorMessage") or EVENT_STATUS.get(op.get("status"), "Failed")
            rec["status"] = EVENT_STATUS.get(op.get("status"), "Failed")
            rec["error"] = msg
            rec["completed_at"] = now_iso()
            self._record_event(name, "failed", msg, op)
            return False
        self._record_event(name, "started", "Reindex event created", op)

        error_count = 0
        while True:
            with self._lock:
                hard_stop = self._stop
            if hard_stop:
                try:
                    cancel_reindex(self._session, self.kibana_url, name)
                except Exception as exc:
                    self._record_event(name, "cancel-failed", str(exc))
                rec["status"] = "Cancelled"
                rec["completed_at"] = now_iso()
                self._record_event(name, "cancelled", "Cancelled via hard stop")
                return False

            try:
                op = get_reindex_status(self._session, self.kibana_url, name)
            except Exception as exc:
                error_count += 1
                if error_count > 5:
                    rec["status"] = "Failed"
                    rec["error"] = str(exc)
                    rec["completed_at"] = now_iso()
                    self._record_event(name, "failed", str(exc))
                    return False
                time.sleep(self.poll_interval)
                continue
            error_count = 0

            if op.get("_error"):
                error_count += 1
                if error_count > 5:
                    rec["status"] = "Failed"
                    rec["error"] = op["_error"]
                    rec["completed_at"] = now_iso()
                    self._record_event(name, "failed", op["_error"])
                    return False
                time.sleep(self.poll_interval)
                continue

            status = op.get("status")
            rec["event"] = op
            if op.get("reindexTaskId"):
                rec["task_id"] = op.get("reindexTaskId")
            new_last = op.get("lastCompletedStep")
            if new_last is not None and new_last != rec.get("step"):
                self._record_event(name, f"step-{new_last}",
                                   _step_milestone_message(name, op.get("newIndexName"), new_last), op)
            rec["step"] = new_last
            rec["step_name"] = step_code_to_name(new_last)
            rec["new_index"] = op.get("newIndexName")
            if op.get("reindexTaskPercComplete") is not None:
                rec["progress"] = round(float(op["reindexTaskPercComplete"]), 1)

            task_id = rec.get("task_id")
            td = None
            if task_id:
                td = get_es_task_status(self._session, self.es_url, task_id)
                if td:
                    rec["task_details"] = td
                    if rec["progress"] is None and td.get("percent") is not None:
                        rec["progress"] = td["percent"]
                    if td.get("docs_total") is not None:
                        rec["docs"] = f"{td['docs_created']}/{td['docs_total']}"

            if status == 1:
                rec["status"] = "Completed"
                rec["step_name"] = "Completed"
                rec["progress"] = 100.0
                rec["completed_at"] = now_iso()
                self._record_event(name, "completed",
                                   f"Reindex completed ({rec['new_index'] or expected_new_index(name)})", op)
                return True
            if status in TERMINAL_FAILED:
                msg = op.get("errorMessage") or EVENT_STATUS.get(status, f"Status {status}")
                rec["status"] = EVENT_STATUS.get(status, "Failed")
                rec["error"] = msg
                rec["completed_at"] = now_iso()
                self._record_event(name, "failed", f"{EVENT_STATUS.get(status, 'Failed')}: {msg}", op)
                return False

            # Reality check: the UA op can stay at status 0 even though ES has
            # finished (stuck reindexOp saved-object, Kibana restart, resumed op).
            # Once the reindex task completed (step >= 50) or the ES task is
            # reported complete, confirm against ES: after the UA swaps the alias
            # and deletes the original, the original name resolves to the
            # reindexed-v8-… index, so version.created >= 8.0 proves completion.
            last = _to_int(new_last) or 0
            es_done = td is not None and td.get("completed")
            if (last >= 50 or es_done) and status != 1 and status not in TERMINAL_FAILED:
                if time.time() - (rec.get("_last_check_at") or 0) >= max(4.0, self.poll_interval * 3):
                    rec["_last_check_at"] = time.time()
                    vr = validate_index(self._session, self.es_url, name)
                    if vr and vr.get("ok"):
                        rec["status"] = "Completed"
                        rec["step_name"] = "Completed"
                        rec["progress"] = 100.0
                        rec["new_index"] = op.get("newIndexName") or expected_new_index(name)
                        rec["version_created"] = vr.get("version_created")
                        if vr.get("docs_count") is not None:
                            rec["docs"] = str(vr["docs_count"])
                        rec["completed_at"] = now_iso()
                        self._record_event(name, "completed",
                                           f"Reindex completed — confirmed via ES alias "
                                           f"({rec['new_index']}, {rec['docs'] or '?'} docs)", op)
                        return True
            time.sleep(self.poll_interval)

    def _process_stream(self, stream_name):
        rec = self._tasks.get(stream_name)
        stream = self._streams.get(stream_name)
        if rec is None or stream is None:
            return
        self._active = stream_name
        try:
            old_backing = stream.get("old_backing", [])
            write = stream.get("write_index")
            if write and write in old_backing:
                rec["status"] = "Rolling Over"
                rec["step_name"] = "Roll over write index (created before 8.0)"
                try:
                    res = rollover_stream(self._session, self.es_url, stream_name)
                except Exception as exc:
                    res = {"_error": str(exc)}
                if res.get("_error"):
                    rec["status"] = "Failed"
                    rec["error"] = res["_error"]
                    rec["completed_at"] = now_iso()
                    self._record_event(stream_name, "write-index-rollover-failed", res["_error"])
                    with self._lock:
                        self._halt = True
                    return
                self._record_event(stream_name, "rolled-over", f"Rolled over write index {write}")

            if not old_backing:
                rec["status"] = "Completed"
                rec["step_name"] = "Completed"
                rec["progress"] = 100.0
                rec["completed_at"] = now_iso()
                self._record_event(stream_name, "completed", "No backing indices needed reindexing")
                return

            rec["status"] = "Running"
            rec["started_at"] = now_iso()

            start = migrate_reindex_stream(self._session, self.es_url, stream_name)
            if start.get("_error"):
                msg = start["_error"]
                rec["status"] = "Failed"
                rec["error"] = msg
                rec["completed_at"] = now_iso()
                self._record_event(stream_name, "migration-start-failed", msg)
                with self._lock:
                    self._halt = True
                return
            self._record_event(stream_name, "migration-started",
                               f"ES migration reindex started for data stream {stream_name}", start)

            for b in old_backing:
                brec = self._tasks.setdefault(b, self._new_record(b, "index", stream_name))
                brec["stream"] = stream_name
                brec["status"] = "In progress"
                brec["step_name"] = "ES migration reindex (whole data stream)"

            total_to_upgrade = None
            error_count = 0
            while True:
                with self._lock:
                    hard_stop = self._stop
                if hard_stop:
                    try:
                        cancel_migrate_reindex(self._session, self.es_url, stream_name)
                    except Exception as exc:
                        self._record_event(stream_name, "cancel-failed", str(exc))
                    rec["status"] = "Stopped"
                    rec["completed_at"] = now_iso()
                    self._record_event(stream_name, "stopped",
                                       "Data-stream migration cancelled via hard stop")
                    return

                st = get_migrate_reindex_status(self._session, self.es_url, stream_name)
                if st.get("_not_found"):
                    error_count += 1
                    if error_count > 30:
                        rec["status"] = "Failed"
                        rec["error"] = ("Migration task status not found — the task did not "
                                        "start or its status already expired")
                        rec["completed_at"] = now_iso()
                        self._record_event(stream_name, "migration-status-lost", rec["error"])
                        with self._lock:
                            self._halt = True
                        return
                    time.sleep(self.poll_interval)
                    continue
                if st.get("_error"):
                    error_count += 1
                    if error_count > 5:
                        rec["status"] = "Failed"
                        rec["error"] = st["_error"]
                        rec["completed_at"] = now_iso()
                        self._record_event(stream_name, "migration-status-failed", st["_error"])
                        with self._lock:
                            self._halt = True
                        return
                    time.sleep(self.poll_interval)
                    continue
                error_count = 0

                total = _to_int(st.get("total_indices_requiring_upgrade"))
                if total is not None:
                    total_to_upgrade = total
                total = total_to_upgrade or 0
                successes = _to_int(st.get("successes")) or 0
                in_prog = (st.get("in_progress") or [{}])[0] if st.get("in_progress") else None
                if total:
                    progress = successes / total * 100
                    if in_prog:
                        t = _to_int(in_prog.get("total_doc_count"))
                        d = _to_int(in_prog.get("reindexed_doc_count"))
                        if t:
                            progress = (successes + (d or 0) / t) / total * 100
                    rec["progress"] = round(progress, 1)
                if in_prog:
                    t = _to_int(in_prog.get("total_doc_count"))
                    d = _to_int(in_prog.get("reindexed_doc_count"))
                    rec["docs"] = f"{d or 0}/{t or 0}" if t else None
                    rec["step_name"] = (f"Reindexing {in_prog.get('index')} "
                                        f"({successes}/{total} backing index(es) done)")
                else:
                    rec["step_name"] = (f"ES migration reindex running — "
                                        f"{successes}/{total} backing index(es) done")

                errors = st.get("errors") or []
                if st.get("complete"):
                    if errors:
                        msg = "; ".join(
                            f"{e.get('index')}: {e.get('message')}" for e in errors
                            if isinstance(e, dict))
                        rec["status"] = "Failed"
                        rec["error"] = msg
                        rec["completed_at"] = now_iso()
                        self._record_event(stream_name, "migration-failed", msg, st)
                        for b in old_backing:
                            brec = self._tasks.get(b)
                            if brec:
                                brec["status"] = "Failed"
                                brec["error"] = msg
                                brec["step_name"] = "Failed"
                        with self._lock:
                            self._halt = True
                        return
                    rec["status"] = "Completed"
                    rec["step_name"] = "Completed"
                    rec["progress"] = 100.0
                    rec["completed_at"] = now_iso()
                    current = get_data_stream_indices(self._session, self.es_url, stream_name)
                    mapping = _map_migrated_backing(old_backing, current)
                    for b in old_backing:
                        brec = self._tasks.get(b)
                        if brec is None:
                            brec = self._tasks.setdefault(
                                b, self._new_record(b, "index", stream_name))
                            brec["stream"] = stream_name
                        brec["status"] = "Completed"
                        brec["step_name"] = "Completed"
                        brec["progress"] = 100.0
                        brec["new_index"] = mapping.get(b) or _migration_new_name(b)
                        brec["completed_at"] = now_iso()
                        self._record_event(b, "completed",
                                           f"Backing index migrated to {brec['new_index']}")
                    self._record_event(stream_name, "completed",
                                       f"All {len(old_backing)} backing index(es) migrated", st)
                    if self.pause_after_each:
                        with self._lock:
                            self._halt = True
                    return

                if st.get("exception"):
                    rec["status"] = "Failed"
                    rec["error"] = st["exception"]
                    rec["completed_at"] = now_iso()
                    self._record_event(stream_name, "migration-failed", st["exception"], st)
                    with self._lock:
                        self._halt = True
                    return

                time.sleep(self.poll_interval)
        finally:
            self._active = None

    def _record_api_call(self, entry: dict):
        with self._lock:
            self._api_log.append(entry)
            if len(self._api_log) > self.max_events:
                self._api_log = self._api_log[-self.max_events:]

    def _record_event(self, name, event_type, message, op=None):
        with self._lock:
            self._events.append({
                "ts": now_iso(), "name": name, "type": event_type,
                "message": message, "event": op,
            })
            if len(self._events) > self.max_events:
                self._events = self._events[-self.max_events:]


# ── Excel tracker ────────────────────────────────────────────────────────────

def build_tracker_excel(state: dict) -> bytes:
    """Four-sheet workbook: Indices, Data Streams, Events, API Audit."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws = wb.active
    ws.title = "Indices"
    ws.append(["Index", "Type", "Stream", "Status", "New Index", "Alias",
               "Version Created", "Docs", "Progress %", "Step Code", "Step Name",
               "Task ID", "ES Docs Created", "ES Docs Total", "ES Docs Updated",
               "ES Docs Deleted", "ES Run (ms)", "Started At", "Completed At", "Error"])
    for name, r in sorted(state.get("tasks", {}).items()):
        if r.get("kind") == "stream":
            continue
        td = r.get("task_details") or {}
        ws.append([
            name, r.get("kind"), r.get("stream") or "", r.get("status") or "",
            r.get("new_index") or expected_new_index(name), r.get("alias") or name,
            r.get("version_created") or "", r.get("docs") or "",
            r.get("progress") if r.get("progress") is not None else "",
            r.get("step") or "", r.get("step_name") or "",
            r.get("task_id") or "", td.get("docs_created") or "",
            td.get("docs_total") or "", td.get("docs_updated") or "",
            td.get("docs_deleted") or "", td.get("running_ms") or "",
            r.get("started_at") or "", r.get("completed_at") or "",
            (r.get("error") or "")[:300],
        ])

    ws2 = wb.create_sheet("Data Streams")
    ws2.append(["Data Stream", "Status", "Current Step", "Progress %", "Write Index",
                "Old Backing", "Backing Count", "Started At", "Completed At", "Error"])
    for name, r in sorted(state.get("tasks", {}).items()):
        if r.get("kind") != "stream":
            continue
        ws2.append([
            name, r.get("status") or "", r.get("step_name") or "", r.get("progress") or "",
            r.get("write_index") or "", r.get("old_backing_count") or 0,
            r.get("backing_count") or 0, r.get("started_at") or "",
            r.get("completed_at") or "", (r.get("error") or "")[:200],
        ])

    ws3 = wb.create_sheet("Events")
    ws3.append(["Timestamp", "Index / Stream", "Type", "Message"])
    for e in reversed(state.get("events", [])):
        ws3.append([e.get("ts"), e.get("name") or "", e.get("type") or "", e.get("message") or ""])

    ws4 = wb.create_sheet("API Audit")
    ws4.append(["Timestamp", "Method", "URL", "Status", "Request Body", "Response"])
    for e in reversed(state.get("api_log", [])):
        ws4.append([e.get("ts"), e.get("method"), e.get("url"), e.get("status"),
                    (e.get("req_body") or "")[:400], (e.get("resp_body") or "")[:400]])

    for sheet, widths in (
        (ws, [42, 8, 24, 14, 26, 24, 14, 12, 10, 10, 34, 26, 12, 12, 12, 12, 12, 22, 22, 60]),
        (ws2, [42, 14, 34, 12, 24, 12, 14, 22, 22, 60]),
        (ws3, [22, 42, 20, 60]),
        (ws4, [22, 8, 60, 10, 50, 80]),
    ):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
